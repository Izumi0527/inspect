"""
日志导出服务

提供日志数据的导出功能，支持CSV、Excel等格式
"""
import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional, BinaryIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import structlog

from src.models.device_log import LogLevel, LogFacility

logger = structlog.get_logger()


class LogExportService:
    """日志导出服务"""
    
    def __init__(self):
        # 日志级别颜色映射
        self.level_colors = {
            LogLevel.CRITICAL.value: "FF0000",  # 红色
            LogLevel.ERROR.value: "FF6600",     # 橙红色
            LogLevel.WARNING.value: "FFCC00",   # 黄色
            LogLevel.INFO.value: "0066CC",      # 蓝色
            LogLevel.DEBUG.value: "999999"      # 灰色
        }
    
    def export_to_csv(
        self, 
        logs: List[Dict[str, Any]], 
        include_raw: bool = False
    ) -> str:
        """导出日志到CSV格式
        
        Args:
            logs: 日志数据列表
            include_raw: 是否包含原始消息
            
        Returns:
            str: CSV格式的字符串
        """
        if not logs:
            return ""
        
        output = io.StringIO()
        
        # 定义CSV字段
        fieldnames = [
            'ID', '设备ID', '日志级别', '设施类型', '来源', 
            '消息', '源IP', '源进程', '日志时间', '采集时间'
        ]
        
        if include_raw:
            fieldnames.append('原始消息')
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # 写入数据
        for log in logs:
            row = {
                'ID': log.get('id', ''),
                '设备ID': log.get('device_id', ''),
                '日志级别': self._translate_level(log.get('level', '')),
                '设施类型': self._translate_facility(log.get('facility', '')),
                '来源': self._translate_source(log.get('source', '')),
                '消息': log.get('message', ''),
                '源IP': log.get('source_ip', ''),
                '源进程': log.get('source_process', ''),
                '日志时间': self._format_datetime(log.get('log_timestamp')),
                '采集时间': self._format_datetime(log.get('collected_at'))
            }
            
            if include_raw:
                row['原始消息'] = log.get('raw_message', '')
            
            writer.writerow(row)
        
        return output.getvalue()
    
    def export_to_excel(
        self, 
        logs: List[Dict[str, Any]], 
        include_raw: bool = False,
        device_name: Optional[str] = None
    ) -> bytes:
        """导出日志到Excel格式
        
        Args:
            logs: 日志数据列表
            include_raw: 是否包含原始消息
            device_name: 设备名称（用于工作表标题）
            
        Returns:
            bytes: Excel文件的二进制数据
        """
        wb = Workbook()
        ws = wb.active
        
        # 设置工作表名称
        sheet_name = f"设备日志_{device_name}" if device_name else "设备日志"
        ws.title = sheet_name[:31]  # Excel工作表名称限制31个字符
        
        # 定义表头
        headers = [
            'ID', '设备ID', '日志级别', '设施类型', '来源', 
            '消息', '源IP', '源进程', '日志时间', '采集时间'
        ]
        
        if include_raw:
            headers.append('原始消息')
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, log in enumerate(logs, 2):
            data = [
                log.get('id', ''),
                log.get('device_id', ''),
                self._translate_level(log.get('level', '')),
                self._translate_facility(log.get('facility', '')),
                self._translate_source(log.get('source', '')),
                log.get('message', ''),
                log.get('source_ip', ''),
                log.get('source_process', ''),
                self._format_datetime(log.get('log_timestamp')),
                self._format_datetime(log.get('collected_at'))
            ]
            
            if include_raw:
                data.append(log.get('raw_message', ''))
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # 根据日志级别设置颜色
                if col == 3:  # 日志级别列
                    level = log.get('level', '')
                    if level in self.level_colors:
                        cell.fill = PatternFill(
                            start_color=self.level_colors[level],
                            end_color=self.level_colors[level],
                            fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF" if level in ['critical', 'error'] else "000000")
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 添加筛选器
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(logs) + 1}"
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_statistics_to_excel(
        self, 
        stats: Dict[str, Any], 
        logs: List[Dict[str, Any]]
    ) -> bytes:
        """导出日志统计到Excel
        
        Args:
            stats: 统计数据
            logs: 日志数据
            
        Returns:
            bytes: Excel文件的二进制数据
        """
        wb = Workbook()
        
        # 创建统计工作表
        stats_ws = wb.active
        stats_ws.title = "统计信息"
        
        # 写入统计信息
        self._write_statistics_sheet(stats_ws, stats)
        
        # 创建日志详情工作表
        if logs:
            logs_ws = wb.create_sheet("日志详情")
            self._write_logs_sheet(logs_ws, logs)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _write_statistics_sheet(self, ws, stats: Dict[str, Any]):
        """写入统计信息工作表"""
        # 标题
        ws.cell(row=1, column=1, value="日志统计报告").font = Font(size=16, bold=True)
        ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        row = 4
        
        # 总体统计
        ws.cell(row=row, column=1, value="总体统计").font = Font(size=14, bold=True)
        row += 1
        ws.cell(row=row, column=1, value="总日志数")
        ws.cell(row=row, column=2, value=stats.get('total_logs', 0))
        row += 1
        ws.cell(row=row, column=1, value="统计时间范围")
        ws.cell(row=row, column=2, value=f"{stats.get('time_range_hours', 24)} 小时")
        row += 2
        
        # 按级别统计
        by_level = stats.get('by_level', {})
        if by_level:
            ws.cell(row=row, column=1, value="按级别统计").font = Font(size=14, bold=True)
            row += 1
            ws.cell(row=row, column=1, value="级别").font = Font(bold=True)
            ws.cell(row=row, column=2, value="数量").font = Font(bold=True)
            row += 1
            
            for level, count in by_level.items():
                ws.cell(row=row, column=1, value=self._translate_level(level))
                ws.cell(row=row, column=2, value=count)
                row += 1
            row += 1
        
        # 按设施统计
        by_facility = stats.get('by_facility', {})
        if by_facility:
            ws.cell(row=row, column=1, value="按设施统计").font = Font(size=14, bold=True)
            row += 1
            ws.cell(row=row, column=1, value="设施").font = Font(bold=True)
            ws.cell(row=row, column=2, value="数量").font = Font(bold=True)
            row += 1
            
            for facility, count in by_facility.items():
                ws.cell(row=row, column=1, value=self._translate_facility(facility))
                ws.cell(row=row, column=2, value=count)
                row += 1
    
    def _write_logs_sheet(self, ws, logs: List[Dict[str, Any]]):
        """写入日志详情工作表"""
        headers = ['ID', '设备ID', '级别', '设施', '消息', '时间']
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # 写入数据
        for row_idx, log in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=log.get('id', ''))
            ws.cell(row=row_idx, column=2, value=log.get('device_id', ''))
            ws.cell(row=row_idx, column=3, value=self._translate_level(log.get('level', '')))
            ws.cell(row=row_idx, column=4, value=self._translate_facility(log.get('facility', '')))
            ws.cell(row=row_idx, column=5, value=log.get('message', ''))
            ws.cell(row=row_idx, column=6, value=self._format_datetime(log.get('log_timestamp')))
    
    def _translate_level(self, level: str) -> str:
        """翻译日志级别"""
        translations = {
            'critical': '严重',
            'error': '错误',
            'warning': '警告',
            'info': '信息',
            'debug': '调试'
        }
        return translations.get(level.lower(), level)
    
    def _translate_facility(self, facility: str) -> str:
        """翻译设施类型"""
        translations = {
            'system': '系统',
            'interface': '接口',
            'security': '安全',
            'routing': '路由',
            'switching': '交换',
            'snmp': 'SNMP',
            'ssh': 'SSH'
        }
        return translations.get(facility.lower(), facility)
    
    def _translate_source(self, source: str) -> str:
        """翻译日志来源"""
        translations = {
            'ssh': 'SSH采集',
            'snmp': 'SNMP采集',
            'syslog': 'Syslog接收',
            'trap': 'SNMP Trap',
            'manual': '手动录入'
        }
        return translations.get(source.lower(), source)
    
    def _format_datetime(self, dt) -> str:
        """格式化日期时间"""
        if dt is None:
            return ""
        
        if isinstance(dt, str):
            try:
                dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
            except:
                return dt
        
        if isinstance(dt, datetime):
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        
        return str(dt)
    
    def get_export_filename(
        self, 
        format_type: str, 
        device_name: Optional[str] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None
    ) -> str:
        """生成导出文件名
        
        Args:
            format_type: 文件格式 (csv, excel)
            device_name: 设备名称
            start_time: 开始时间
            end_time: 结束时间
            
        Returns:
            str: 文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 基础文件名
        if device_name:
            base_name = f"设备日志_{device_name}_{timestamp}"
        else:
            base_name = f"设备日志_{timestamp}"
        
        # 添加时间范围
        if start_time and end_time:
            date_range = f"{start_time.strftime('%m%d')}_{end_time.strftime('%m%d')}"
            base_name = f"设备日志_{date_range}_{timestamp}"
        
        # 添加扩展名
        if format_type.lower() == 'csv':
            return f"{base_name}.csv"
        elif format_type.lower() in ['excel', 'xlsx']:
            return f"{base_name}.xlsx"
        else:
            return f"{base_name}.txt"